import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

#setando a aparencia padrao do sistema
ctk. set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
           self.title("Sistema de gestão de clientes")
           self.geometry("600x400")

    def appearence(self):
           self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000","#fff"]).place(x=35, y=320)
           self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark"], command=self.change_apm).place(x=30, y=350)


    def todo_sistema(self):
           frame = ctk.CTkFrame(self, width=600, height=40,corner_radius=0, bg_color="teal", fg_color="teal")
           frame.place(x=0, y=3)
           title = ctk.CTkLabel(frame, text="Sistema de Gestão de Clientes", font=("century Gothic bold", 24), text_color="#fff").place(x=120, y=10)

           span = ctk.CTkLabel(self, text="Por Favor, preencha todos os campos do formulário!", font=("century Gothic bold", 16), text_color=["#000","#fff"]).place(x=40, y=60)

           ficheiro = pathlib.Path("Clientes.xlsx")

           if ficheiro.exists():
              pass
           else:
               ficheiro=Workbook()
               folha=ficheiro.active
               folha['A1'] = "Nome completo"
               folha['B1'] = "Contato"
               folha['C1'] = "Idade"
               folha['D1'] = "Gênero"
               folha['E1'] = "Endereço"
               folha['F1'] = "Observações"

               ficheiro.save("Clientes.xlsx")

           def submit():

               #Pegando os dados ds entry
               name = name_value.get()
               contact = contact_value.get()
               age = age_value.get()
               gender = gender_combobox.get()
               address = address_value.get()
               options = options_entry.get(0.0, END)

               if (name =="" or contact=="" or age=="" or address==""):
                    messagebox.showerror('Sistema', "ERRO!\nPor favor preencha todos os campos!")

               ficheiro = openpyxl.load_workbook('Clientes.xlsx')
               folha = ficheiro.active
               folha.cell(column=1, row=folha.max_row+1, value=name)
               folha.cell(column=2, row=folha.max_row, value=contact)
               folha.cell(column=3, row=folha.max_row, value=age)
               folha.cell(column=4, row=folha.max_row, value=gender)
               folha.cell(column=5, row=folha.max_row, value=address)
               folha.cell(column=6, row=folha.max_row, value=options)

               ficheiro.save(r"Clientes.xlsx")
               messagebox.showinfo("Systema", "Dados salvos com sucesso!")



           def clear():
               name_value.set("")
               contact_value.set("")
               age_value.set("")
               address_value.set("")
               options_entry.delete(0.0, END)
           


           #texts variable
           name_value = StringVar()
           contact_value = StringVar()
           age_value = StringVar()
           address_value = StringVar()

           #entrys
           name_entry = ctk.CTkEntry(self, width=230, textvariable=name_value, font=("centry Gohtic", 16), fg_color="transparent")
           contact_entry = ctk.CTkEntry(self, width=180, textvariable=contact_value, font=("centry Gohtic", 16), fg_color="transparent")
           age_entry = ctk.CTkEntry(self, width=130, textvariable=age_value, font=("centry Gohtic", 16), fg_color="transparent")
           address_entry = ctk.CTkEntry(self, width=180, textvariable=address_value, font=("centry Gohtic", 16), fg_color="transparent")


           #Combobox
           gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("centry Gohtic", 14))
           gender_combobox.set("Masculino")

           #Entrada de observações
           options_entry = ctk.CTkTextbox(self, width=400, height=70, font=("arial", 18), border_color="#aaa", border_width=1, fg_color="transparent")

           #Labels
           lb_name = ctk.CTkLabel(self, text="Nome completo", font=("Century Gothic bold", 14), text_color=["#000","#fff"])
           lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold", 14), text_color=["#000","#fff"])
           lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold", 14), text_color=["#000","#fff"])
           lb_gender = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic bold", 14), text_color=["#000","#fff"])
           lb_address = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold", 14), text_color=["#000","#fff"])
           lb_options = ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold", 14), text_color=["#000","#fff"])

           btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=200, y=320)
           btn_submit = ctk.CTkButton(self, text="Limpar dados".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=400, y=320)

           #posicionando os elemtentos na janela
           lb_name.place(x=40, y=100)
           name_entry.place(x=40, y=130)

           lb_contact.place(x=380, y=100)
           contact_entry.place(x=380, y=130)

           lb_age.place(x=250, y=160)
           age_entry.place(x=250, y=190)

           lb_gender.place(x=400, y=160)
           gender_combobox.place(x=400, y=190)

           lb_address.place(x=40, y=160)
           address_entry.place(x=40, y=190)

           lb_options.place(x=40, y=220)
           options_entry.place(x=135, y=230)


    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__=="__main__":
   app = App()
   app.mainloop()